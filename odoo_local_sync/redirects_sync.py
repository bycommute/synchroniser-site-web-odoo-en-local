"""
Synchronisation des redirections Odoo (website.rewrite) depuis/vers Excel.
"""
from __future__ import annotations

from dataclasses import dataclass
import os
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple
from urllib.parse import urlsplit

from .config import WEBSITE_ID
from .odoo_client import get_client


MANAGED_NAME_PREFIX = "[odoo-local-sync]"
LOCAL_REDIRECT_HOSTS = {
    host.strip().lower()
    for host in os.getenv("ODOO_REDIRECT_LOCAL_HOSTS", "").split(",")
    if host.strip()
}


@dataclass
class RedirectRow:
    row_number: int
    url_from: str
    url_to: str


class RedirectSyncError(Exception):
    pass


def _normalize_path(raw: str, *, for_source: bool) -> str:
    value = (raw or "").strip()
    if not value:
        return ""

    parts = urlsplit(value)
    has_host = bool(parts.netloc)

    if has_host:
        host = (parts.netloc or "").lower()
        if host in LOCAL_REDIRECT_HOSTS or any(host.endswith("." + allowed) for allowed in LOCAL_REDIRECT_HOSTS):
            path = parts.path or "/"
            if not path.startswith("/"):
                path = "/" + path
            out = path
            if parts.query:
                out += "?" + parts.query
            return out
        # URL externe: on la garde telle quelle pour destination, interdite pour source.
        if for_source:
            raise RedirectSyncError(f"URL source externe non autorisée: {value}")
        return value

    path = value
    if not path.startswith("/"):
        path = "/" + path
    return path


def _chunked(items: List[str], size: int = 100) -> List[List[str]]:
    return [items[i : i + size] for i in range(0, len(items), size)]


class RedirectSyncService:
    def __init__(self):
        self.client = get_client()

    def _read_excel_rows(self, excel_path: Path, sheet_name: str) -> List[RedirectRow]:
        try:
            import openpyxl
        except ImportError as exc:
            raise RedirectSyncError("openpyxl requis pour lire/écrire le fichier Excel.") from exc

        if not excel_path.exists():
            raise RedirectSyncError(f"Fichier Excel introuvable: {excel_path}")

        wb = openpyxl.load_workbook(excel_path, data_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            raise RedirectSyncError(f"Onglet introuvable: {sheet_name}")

        ws = wb[sheet_name]
        header = [str(c.value or "").strip().lower() for c in ws[1]]
        col_old = None
        col_new = None

        for idx, name in enumerate(header, start=1):
            if name in {"ancienne url", "url source", "source", "old url"}:
                col_old = idx
            if name in {"nouvelle url", "url destination", "destination", "new url"}:
                col_new = idx

        # fallback historique ByCommute: col A/B
        if col_old is None:
            col_old = 1
        if col_new is None:
            col_new = 2

        rows: List[RedirectRow] = []
        for row_idx in range(2, ws.max_row + 1):
            raw_old = ws.cell(row=row_idx, column=col_old).value
            raw_new = ws.cell(row=row_idx, column=col_new).value
            if raw_old is None or raw_new is None:
                continue

            try:
                old_path = _normalize_path(str(raw_old), for_source=True)
                new_path = _normalize_path(str(raw_new), for_source=False)
            except RedirectSyncError as exc:
                wb.close()
                raise RedirectSyncError(f"Ligne {row_idx}: {exc}") from exc

            if not old_path or not new_path:
                continue

            rows.append(RedirectRow(row_number=row_idx, url_from=old_path, url_to=new_path))

        wb.close()
        return rows

    def _fetch_existing_by_sources(self, sources: List[str]) -> Dict[str, List[Dict]]:
        existing: Dict[str, List[Dict]] = {}
        if not sources:
            return existing

        fields = ["id", "name", "url_from", "url_to", "redirect_type", "active", "website_id"]
        for part in _chunked(sources, size=100):
            records = self.client.search_read(
                "website.rewrite",
                domain=[["url_from", "in", part]],
                fields=fields,
            )
            for rec in records:
                key = rec.get("url_from")
                if not key:
                    continue
                existing.setdefault(key, []).append(rec)
        return existing

    @staticmethod
    def _is_current_or_global_website(record: Dict) -> bool:
        website = record.get("website_id")
        return (not website) or (
            isinstance(website, (list, tuple)) and website and int(website[0]) == WEBSITE_ID
        )

    @staticmethod
    def _is_managed_record(record: Dict) -> bool:
        name = record.get("name") or ""
        return isinstance(name, str) and name.startswith(MANAGED_NAME_PREFIX)

    def _fetch_existing_filtered(
        self,
        *,
        only_active: Optional[bool] = None,
        managed_only: bool = False,
    ) -> List[Dict]:
        fields = ["id", "name", "url_from", "url_to", "redirect_type", "active", "website_id"]
        records = self.client.search_read("website.rewrite", domain=[], fields=fields)

        filtered: List[Dict] = []
        for rec in records:
            if not self._is_current_or_global_website(rec):
                continue
            if only_active is True and not rec.get("active"):
                continue
            if only_active is False and rec.get("active"):
                continue
            if managed_only and not self._is_managed_record(rec):
                continue
            filtered.append(rec)
        return filtered

    def _pick_best_record(self, records: List[Dict], desired_url_to: str) -> Dict:
        if not records:
            return {}

        def score(rec: Dict) -> Tuple[int, int, int, int]:
            website = rec.get("website_id")
            website_score = 2 if isinstance(website, (list, tuple)) and website and int(website[0]) == WEBSITE_ID else 1 if not website else 0
            active_score = 1 if rec.get("active") else 0
            managed_score = 1 if self._is_managed_record(rec) else 0
            target_score = 1 if rec.get("url_to") == desired_url_to else 0
            return (target_score, website_score, managed_score, active_score)

        return max(records, key=score)

    def push_from_excel(
        self,
        excel_path: Path,
        sheet_name: str = "Redirections",
        dry_run: bool = True,
        prune: bool = False,
        takeover_existing: bool = False,
        prune_unmanaged: bool = False,
    ) -> Dict:
        rows = self._read_excel_rows(excel_path, sheet_name)
        desired: Dict[str, RedirectRow] = {}
        duplicates: List[str] = []
        self_redirects_skipped = 0

        for row in rows:
            # Odoo interdit les auto-redirections (source == destination)
            if row.url_from == row.url_to:
                self_redirects_skipped += 1
                continue
            prev = desired.get(row.url_from)
            if prev and prev.url_to != row.url_to:
                duplicates.append(
                    f"Source en doublon avec destination différente: {row.url_from} (lignes {prev.row_number} et {row.row_number})"
                )
            desired[row.url_from] = row

        if duplicates:
            raise RedirectSyncError("Incohérences Excel:\n- " + "\n- ".join(duplicates))

        sources = sorted(desired.keys())
        existing_map = self._fetch_existing_by_sources(sources)

        to_create: List[Dict] = []
        to_update: List[Tuple[int, Dict]] = []
        to_take_over = 0
        unchanged = 0
        ambiguous = []
        duplicate_deactivations: Set[int] = set()

        for source in sources:
            wanted = desired[source]
            records = existing_map.get(source, [])
            active_records = [rec for rec in records if rec.get("active")]
            if len(active_records) > 1:
                ambiguous.append(source)
            best = self._pick_best_record(records, wanted.url_to)
            if not best:
                to_create.append(
                    {
                        "name": f"{MANAGED_NAME_PREFIX} Redirect {source}",
                        "url_from": source,
                        "url_to": wanted.url_to,
                        "redirect_type": "301",
                        "active": True,
                        "website_id": WEBSITE_ID,
                    }
                )
                continue

            patch: Dict = {}
            desired_name = f"{MANAGED_NAME_PREFIX} Redirect {source}"
            if best.get("url_to") != wanted.url_to:
                patch["url_to"] = wanted.url_to
            if best.get("redirect_type") != "301":
                patch["redirect_type"] = "301"
            if best.get("active") is not True:
                patch["active"] = True
            if takeover_existing:
                website = best.get("website_id")
                current_website_id = website[0] if isinstance(website, (list, tuple)) and website else None
                if best.get("name") != desired_name:
                    patch["name"] = desired_name
                if current_website_id != WEBSITE_ID:
                    patch["website_id"] = WEBSITE_ID
            if patch:
                if takeover_existing and (
                    "name" in patch or "website_id" in patch
                ):
                    to_take_over += 1
                to_update.append((int(best["id"]), patch))
            else:
                unchanged += 1

            if takeover_existing:
                best_id = int(best["id"])
                for rec in active_records:
                    rec_id = int(rec["id"])
                    if rec_id != best_id:
                        duplicate_deactivations.add(rec_id)

        managed_pruned: Set[int] = set()
        if prune:
            managed = self._fetch_existing_filtered(only_active=True, managed_only=True)
            for rec in managed:
                src = rec.get("url_from") or ""
                if src not in desired:
                    managed_pruned.add(int(rec["id"]))

        unmanaged_pruned: Set[int] = set()
        if prune_unmanaged:
            existing_active = self._fetch_existing_filtered(only_active=True, managed_only=False)
            for rec in existing_active:
                if self._is_managed_record(rec):
                    continue
                src = rec.get("url_from") or ""
                if src not in desired:
                    unmanaged_pruned.add(int(rec["id"]))

        to_deactivate = sorted(duplicate_deactivations | managed_pruned | unmanaged_pruned)

        if not dry_run:
            if to_create:
                self.client.execute("website.rewrite", "create", to_create)
            for rec_id, vals in to_update:
                self.client.write("website.rewrite", [rec_id], vals)
            if to_deactivate:
                self.client.write("website.rewrite", to_deactivate, {"active": False})

        return {
            "total_excel_rows": len(rows),
            "unique_sources": len(sources),
            "self_redirects_skipped": self_redirects_skipped,
            "to_create": len(to_create),
            "to_update": len(to_update),
            "to_take_over": to_take_over,
            "unchanged": unchanged,
            "ambiguous_sources": ambiguous,
            "to_deactivate_duplicates": len(duplicate_deactivations),
            "to_deactivate_managed": len(managed_pruned),
            "to_deactivate_unmanaged": len(unmanaged_pruned),
            "to_deactivate": len(to_deactivate),
            "dry_run": dry_run,
        }

    def pull_to_excel(
        self,
        excel_path: Path,
        sheet_name: str = "Redirections Odoo",
        only_active: bool = True,
    ) -> Dict:
        try:
            import openpyxl
        except ImportError as exc:
            raise RedirectSyncError("openpyxl requis pour lire/écrire le fichier Excel.") from exc

        filtered = self._fetch_existing_filtered(only_active=True if only_active else None, managed_only=False)

        filtered.sort(key=lambda r: (r.get("url_from") or ""))

        wb = openpyxl.load_workbook(excel_path) if excel_path.exists() else openpyxl.Workbook()
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            wb.remove(ws)
        ws = wb.create_sheet(sheet_name)

        headers = ["ID Odoo", "Nom", "Ancienne URL", "Nouvelle URL", "Type", "Active", "Website ID"]
        ws.append(headers)
        for rec in filtered:
            website = rec.get("website_id")
            website_id = website[0] if isinstance(website, (list, tuple)) and website else ""
            ws.append(
                [
                    rec.get("id"),
                    rec.get("name") or "",
                    rec.get("url_from") or "",
                    rec.get("url_to") or "",
                    rec.get("redirect_type") or "",
                    bool(rec.get("active")),
                    website_id,
                ]
            )

        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 50
        ws.column_dimensions["D"].width = 70
        ws.column_dimensions["E"].width = 10
        ws.column_dimensions["F"].width = 10
        ws.column_dimensions["G"].width = 10

        wb.save(excel_path)
        wb.close()

        return {
            "count": len(filtered),
            "excel_path": str(excel_path),
            "sheet_name": sheet_name,
        }
